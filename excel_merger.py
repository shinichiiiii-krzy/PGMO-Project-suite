"""
excel_merger.py

Drop-in module for merging multiple Excel sheets, removing exact-duplicate
rows, sorting by LCP + NAP code, and building an AREA / port-count pivot
table -- all written into a single output workbook.

Requires: pandas, openpyxl
    pip install pandas openpyxl

Typical usage
-------------
    from excel_merger import process_workbooks

    result = process_workbooks(
        input_paths=["site_a.xlsx", "site_b.xlsx", "site_c.xlsx"],
        output_path="merged_network_data.xlsx",
    )

    print(result.summary())

If your column names don't exactly match "LCP", "NAP CODE", "AREA", "PORTS"
etc., either pass the exact names explicitly, or let auto-detection run and
check `result.columns_used` to confirm it guessed right.

Columns with a blank/merged header cell in the source file are never
dropped -- they're kept and labeled "Column N (blank header)" so they
stay selectable (see `resolve_header_names`).

Rows removed as exact duplicates aren't just discarded: pass
`duplicates_csv_path`, or leave it on the default, and they're written
to their own CSV next to the output workbook. Pass `log_path` to also
save the full run log to a text file, and/or `log_fn` for a live
progress callback (e.g. wired into a GUI console).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Optional, Union

import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Imported directly (even though nothing below calls it by name) so that
# PyInstaller's static analysis sees it as a real dependency and bundles it
# into the .exe. pandas only reaches xlrd dynamically via engine="xlrd",
# which PyInstaller's import scanner cannot see -- without this line the
# packaged app builds fine, runs fine on .xlsx files, and then fails with
# "install xlrd" the moment someone opens a legacy .xls file.
try:
    import xlrd  # noqa: F401
except ImportError:
    xlrd = None


# --------------------------------------------------------------------------- #
# Column auto-detection
# --------------------------------------------------------------------------- #

_LCP_PATTERNS = [r"^lcp$", r"\blcp\b"]
_NAP_PATTERNS = [r"^nap[\s_-]*code$", r"\bnap\b"]
_AREA_PATTERNS = [r"^area$", r"\barea\b"]
_PORTS_PATTERNS = [
    r"no\.?\s*of\s*ports?\b",   # "NO. OF PORT" / "NO. OF PORTS"
    r"num.*ports?\b",
    r"^ports?$",
    r"port\s*capacity",
    r"port_capacity",
    r"\bport\b",                 # broad fallback -- last resort only
]


def _guess_column(columns: Iterable[str], patterns: list[str]) -> Optional[str]:
    """Return the first column whose name matches any pattern, case-insensitive."""
    cols = list(columns)
    for pattern in patterns:
        regex = re.compile(pattern, re.IGNORECASE)
        for col in cols:
            if regex.search(str(col)):
                return col
    return None


def resolve_header_names(columns: Iterable) -> list[str]:
    """Turn raw pandas column headers into names that are always visible
    and selectable.

    Source workbooks (especially telecom/patching trackers) often have
    merged or blank header cells -- pandas reads those as "" or
    "Unnamed: N". Previously those columns were silently dropped, which
    meant a real data column (e.g. AREA or PORTS) could disappear from
    the picker entirely with no way to select it. Instead, give every
    blank-header column a clear placeholder name based on its position,
    so it always shows up as a choosable option.
    """
    resolved = []
    for i, c in enumerate(columns):
        c_str = str(c).strip()
        if c_str == "" or c_str.lower().startswith("unnamed"):
            resolved.append(f"Column {i + 1} (blank header)")
        else:
            resolved.append(c_str)
    return resolved


def detect_header_row(path: Union[str, Path], sheet_name: Union[str, int] = 0, max_rows_to_check: int = 5) -> int:
    """Some exports have a hidden machine-code row above the real, human-
    readable header (e.g. row 0 = 'lcp_id', row 1 = 'LCP'). Try the first
    few rows as a candidate header and return the 0-based index of whichever
    one matches the most of LCP/NAP/AREA/PORT -- that's almost always the
    real header row."""
    raw = _read_excel(
        path, sheet_name=sheet_name, header=None, nrows=max_rows_to_check, dtype=str,
    )
    best_row, best_score = 0, -1
    for i in range(len(raw)):
        candidate_cols = [str(v) for v in raw.iloc[i].tolist()]
        score = sum(
            1 for patterns in (_LCP_PATTERNS, _NAP_PATTERNS, _AREA_PATTERNS, _PORTS_PATTERNS)
            if _guess_column(candidate_cols, patterns)
        )
        if score > best_score:
            best_score, best_row = score, i
    return best_row


# --------------------------------------------------------------------------- #
# Result container
# --------------------------------------------------------------------------- #

@dataclass
class MergeResult:
    combined_df: pd.DataFrame
    pivot_df: pd.DataFrame
    duplicate_rows_df: pd.DataFrame
    rows_merged: int
    blank_rows_removed: int
    duplicates_removed: int
    rows_final: int
    grand_total_ports: float
    columns_used: dict = field(default_factory=dict)
    output_path: Optional[Path] = None
    duplicates_csv_path: Optional[Path] = None
    log_path: Optional[Path] = None

    def summary(self) -> str:
        lines = [
            f"Files merged into {self.rows_merged} row(s)",
            f"Blank rows removed: {self.blank_rows_removed}",
            f"Duplicates removed: {self.duplicates_removed}",
            f"Final row count: {self.rows_final}",
            f"Areas in pivot: {len(self.pivot_df)}",
            f"Grand total ports: {self.grand_total_ports}",
            f"Columns used: {self.columns_used}",
        ]
        if self.output_path:
            lines.append(f"Written to: {self.output_path}")
        if self.duplicates_csv_path:
            lines.append(f"Removed duplicates saved to: {self.duplicates_csv_path}")
        if self.log_path:
            lines.append(f"Log saved to: {self.log_path}")
        return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Core steps (each usable standalone if you want to wire them differently)
# --------------------------------------------------------------------------- #

def _sniff_excel_kind(path: Union[str, Path]) -> Optional[str]:
    """Look at the file's actual bytes, not its extension, to tell modern
    Excel files from legacy ones. A file renamed or mislabeled with the
    wrong extension (a .xls that's really .xlsx content, or vice versa --
    common after email attachments or old export scripts) still opens
    correctly this way. Returns "xlsx", "xls", or None if unrecognized."""
    try:
        with open(path, "rb") as f:
            header = f.read(8)
    except OSError:
        return None
    if header.startswith(b"PK\x03\x04"):
        return "xlsx"   # modern zip-based container: .xlsx / .xlsm
    if header.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls"    # legacy OLE2 binary format: .xls
    return None


def _engine_for(path: Union[str, Path]) -> Optional[str]:
    """Pick the right pandas engine to open an Excel file, no matter what
    its extension is. Sniffs the real file signature first; only falls
    back to the extension if the signature isn't recognized."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return None          # handled separately, not an Excel engine

    kind = _sniff_excel_kind(path)
    if kind == "xls":
        return "xlrd"
    if kind == "xlsx":
        return "openpyxl"

    # Signature unrecognized (unusual/corrupted header) -- fall back to
    # trusting the extension so we still make a reasonable attempt.
    if ext == ".xls":
        return "xlrd"
    return "openpyxl"        # default: the modern, far more common format


def _read_excel(path: Union[str, Path], **kwargs) -> pd.DataFrame:
    """pd.read_excel wrapper that resolves the engine automatically and
    turns a missing-package failure into an actionable message instead of
    a raw ImportError/BadZipFile traceback."""
    engine = _engine_for(path)
    if engine == "xlrd" and xlrd is None:
        raise ImportError(
            f"'{Path(path).name}' looks like a legacy .xls file, which needs the "
            f"'xlrd' package to open. Install it with:  pip install xlrd\n"
            f"(If you're running the packaged .exe, this build wasn't compiled "
            f"with xlrd bundled -- rebuild after adding xlrd to requirements.txt.)"
        )
    try:
        return pd.read_excel(path, engine=engine, **kwargs)
    except ImportError as e:
        pkg = "xlrd" if engine == "xlrd" else "openpyxl"
        raise ImportError(
            f"Couldn't open '{Path(path).name}' -- the '{pkg}' package is required "
            f"but not available. Install it with:  pip install {pkg}\n"
            f"(Original error: {e})"
        ) from e


def load_and_merge(
    input_paths: Iterable[Union[str, Path]],
    sheet_name: Union[str, int] = 0,
    header_row: Optional[int] = None,
    log_fn: Optional[Callable[[str, Optional[str]], None]] = None,
) -> pd.DataFrame:
    """Read the given sheet (default: first sheet) from every file and
    concatenate them into one DataFrame. Missing columns across files are
    filled with empty strings rather than raising an error.

    header_row: 0-based row index where the real column headers live.
        - None (default): auto-detect the header row SEPARATELY for each
          file (via detect_header_row). Different source exports often
          have their header on a different row -- e.g. one file has a
          hidden machine-code row above the human-readable header and
          another doesn't -- so forcing the same row index onto every
          file silently reads a data row as the header for whichever
          files don't match, which corrupts that file's columns and
          shows up later as a wall of sparse/missing cells after concat.
        - An int: force that same 0-based row index for every file
          (the old behavior). Use this only if you're certain every
          input file's header genuinely sits on the same row.

    log_fn: optional callback(message, tag) for progress/status logging.
    """
    def _log(message: str, tag: Optional[str] = "info") -> None:
        if log_fn:
            log_fn(message, tag)

    frames = []
    for p in input_paths:
        suffix = Path(p).suffix.lower()
        if header_row is None and suffix != ".csv":
            this_header_row = detect_header_row(p, sheet_name=sheet_name)
            _log(
                f"  {Path(p).name}: auto-detected header on row {this_header_row + 1}.",
                "info",
            )
        else:
            this_header_row = header_row if header_row is not None else 0

        if suffix == ".csv":
            df = pd.read_csv(p, dtype=str, keep_default_na=False, header=this_header_row)
        else:
            df = _read_excel(
                p, sheet_name=sheet_name, dtype=str, keep_default_na=False, header=this_header_row,
            )

        # Give every blank/"Unnamed" header a visible placeholder name
        # instead of dropping the column outright -- a real data column
        # can sit under a merged/blank header cell and must stay
        # selectable. Only truly empty columns (blank header AND no data
        # in any row) get dropped, since those are genuine leftover junk
        # columns from the export.
        resolved_names = resolve_header_names(df.columns)
        drop_positions = []
        for i, (orig, resolved) in enumerate(zip(df.columns, resolved_names)):
            was_blank_header = resolved != str(orig).strip()
            if was_blank_header:
                has_data = df.iloc[:, i].astype(str).str.strip().ne("").any()
                if not has_data:
                    drop_positions.append(i)
        df.columns = resolved_names
        if drop_positions:
            dropped_names = [resolved_names[i] for i in drop_positions]
            keep_positions = [i for i in range(len(df.columns)) if i not in drop_positions]
            df = df.iloc[:, keep_positions]
            _log(f"  {Path(p).name}: dropped {len(drop_positions)} fully empty column(s): {dropped_names}", "info")

        _log(f"Loaded {len(df)} row(s) from {Path(p).name} -- columns: {list(df.columns)}", "info")
        frames.append(df)
    if not frames:
        raise ValueError("No input files provided.")
    combined = pd.concat(frames, ignore_index=True, sort=False)
    return combined


def remove_blank_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Drop rows where every column is empty/whitespace-only."""
    is_blank = df.apply(lambda row: all(str(v).strip() == "" for v in row), axis=1)
    cleaned = df[~is_blank].reset_index(drop=True)
    blank_rows_removed = int(is_blank.sum())
    return cleaned, blank_rows_removed


def remove_full_row_duplicates(df: pd.DataFrame) -> tuple[pd.DataFrame, int, pd.DataFrame]:
    """Remove a row only if every column is identical to a previously seen
    row (whitespace-trimmed, case-insensitive). Keeps the first occurrence.

    Returns (deduped_df, duplicates_removed_count, duplicate_rows_df) --
    duplicate_rows_df holds the actual rows that got removed, with a
    leading "Original row #" column (1-based, matching the row's position
    in the merged-but-not-yet-deduped data) so they can be cross-checked
    against the source later.
    """
    normalized = df.apply(lambda col: col.astype(str).str.strip().str.lower())
    is_dup = normalized.duplicated(keep="first")

    deduped = df[~is_dup].reset_index(drop=True)
    duplicates_removed = int(is_dup.sum())

    duplicate_rows = df[is_dup].copy()
    duplicate_rows.insert(0, "Original row #", duplicate_rows.index + 1)
    duplicate_rows = duplicate_rows.reset_index(drop=True)

    return deduped, duplicates_removed, duplicate_rows


def write_removed_rows_csv(duplicate_rows: pd.DataFrame, output_path: Union[str, Path]) -> Path:
    """Write the rows that were removed as exact duplicates to their own
    CSV file, so nothing that gets dropped from the merge is ever lost --
    it's just filed separately for review."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    duplicate_rows.to_csv(output_path, index=False)
    return output_path


def _natural_key(val) -> tuple:
    """Split a value into text/number chunks so 'LCP-2' sorts before
    'LCP-10'. Returns a tuple of (type_flag, value) pairs so mixed
    text/number chunks always compare safely."""
    parts = re.split(r"(\d+)", str(val))
    key = []
    for p in parts:
        if p == "":
            continue
        if p.isdigit():
            key.append((0, int(p)))  # numbers sort before text at same position
        else:
            key.append((1, p.lower()))
    return tuple(key)


def sort_by_lcp_nap(df: pd.DataFrame, lcp_col: str, nap_col: str) -> pd.DataFrame:
    """Sort by LCP first, then NAP code, using a natural sort (so 'LCP-2'
    sorts before 'LCP-10')."""
    order = sorted(
        range(len(df)),
        key=lambda i: (_natural_key(df.iloc[i][lcp_col]), _natural_key(df.iloc[i][nap_col])),
    )
    return df.iloc[order].reset_index(drop=True)


def build_area_pivot(df: pd.DataFrame, area_col: str, ports_col: str) -> tuple[pd.DataFrame, float]:
    """Group by AREA, sum the ports column, sort areas alphabetically, and
    append a GRAND TOTAL row. Non-numeric port values are treated as 0."""
    working = df.copy()
    working[area_col] = working[area_col].astype(str).str.strip()
    working.loc[working[area_col] == "", area_col] = "(blank)"
    working["_ports_numeric"] = pd.to_numeric(
        working[ports_col].astype(str).str.replace(r"[^0-9.\-]", "", regex=True),
        errors="coerce",
    ).fillna(0)

    pivot = (
        working.groupby(area_col, as_index=False)["_ports_numeric"]
        .sum()
        .rename(columns={area_col: "AREA", "_ports_numeric": "TOTAL PORTS"})
        .sort_values("AREA", kind="stable")
        .reset_index(drop=True)
    )
    grand_total = pivot["TOTAL PORTS"].sum()

    total_row = pd.DataFrame([{"AREA": "GRAND TOTAL", "TOTAL PORTS": grand_total}])
    pivot_with_total = pd.concat([pivot, total_row], ignore_index=True)
    return pivot_with_total, grand_total


# --------------------------------------------------------------------------- #
# Presentation formatting
# --------------------------------------------------------------------------- #

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=10)
_TOTAL_FONT = Font(name="Arial", size=10, bold=True)
_TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOP_DOUBLE = Border(left=_THIN, right=_THIN, top=Side(style="double", color="1F3864"), bottom=_THIN)


def _autosize_columns(ws, df: pd.DataFrame, max_width: int = 45, min_width: int = 9):
    for i, col in enumerate(df.columns, start=1):
        header_len = len(str(col)) + 4          # bold header text needs extra room
        data_len = max([len(str(v)) for v in df[col].astype(str).tolist()], default=0) + 2
        width = max(header_len, data_len)
        ws.column_dimensions[get_column_letter(i)].width = max(min_width, min(max_width, width))


def _style_data_sheet(ws, df: pd.DataFrame, total_row_label: Optional[str] = None):
    """Bold header row with fill, thin borders, zebra striping, frozen header,
    and an autofilter. If total_row_label is given, the matching last row
    (e.g. 'GRAND TOTAL') gets bold text and a highlight instead of banding."""
    n_rows, n_cols = df.shape

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 22

    is_total_row = False
    for r in range(2, n_rows + 2):
        first_cell_value = ws.cell(row=r, column=1).value
        is_total_row = bool(total_row_label) and str(first_cell_value).strip().upper() == total_row_label.upper()
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _TOP_DOUBLE if is_total_row else _BORDER
            if is_total_row:
                cell.font = _TOTAL_FONT
                cell.fill = _TOTAL_FILL
            else:
                cell.font = _BODY_FONT
                if (r % 2) == 1:
                    cell.fill = _ZEBRA_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws, df)


def write_formatted_workbook(
    combined_df: pd.DataFrame,
    pivot_df: pd.DataFrame,
    output_path: Union[str, Path],
    combined_sheet_name: str = "Combined",
    pivot_sheet_name: str = "Pivot Table",
):
    """Write combined + pivot sheets with presentation formatting: bold
    header row, borders, zebra striping, frozen header, autofilter,
    auto-sized columns, and a highlighted grand-total row on the pivot."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name=combined_sheet_name, index=False)
        pivot_df.to_excel(writer, sheet_name=pivot_sheet_name, index=False)
        wb = writer.book
        _style_data_sheet(wb[combined_sheet_name], combined_df)
        _style_data_sheet(wb[pivot_sheet_name], pivot_df, total_row_label="GRAND TOTAL")


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def process_workbooks(
    input_paths: Iterable[Union[str, Path]],
    output_path: Optional[Union[str, Path]] = None,
    sheet_name: Union[str, int] = 0,
    header_row: Optional[int] = None,
    lcp_col: Optional[str] = None,
    nap_col: Optional[str] = None,
    area_col: Optional[str] = None,
    ports_col: Optional[str] = None,
    combined_sheet_name: str = "Combined",
    pivot_sheet_name: str = "Pivot Table",
    duplicates_csv_path: Optional[Union[str, Path]] = None,
    write_duplicates_csv: bool = True,
    log_path: Optional[Union[str, Path]] = None,
    log_fn: Optional[Callable[[str, Optional[str]], None]] = None,
) -> MergeResult:
    """
    Run the full pipeline: merge -> dedupe (full row) -> sort (LCP, NAP) ->
    pivot (AREA sum of PORTS). Writes an .xlsx with two sheets if
    output_path is given, and always returns a MergeResult with the
    DataFrames and stats.

    Pass lcp_col / nap_col / area_col / ports_col explicitly if you don't
    want auto-detection (recommended for a production integration, so a
    renamed column can't silently change what gets sorted/pivoted).

    header_row: leave as None (default) to auto-detect the header row
    separately for each input file -- recommended, since mixed-source
    files often don't all have their header on the same row. Pass an int
    to force that same row for every file instead.

    duplicates_csv_path: where to save the rows removed as exact
    duplicates. If omitted and write_duplicates_csv is True (default),
    it's derived automatically from output_path as
    "<output name>_duplicates_removed.csv" in the same folder.

    log_path: if given, every log_fn message from this run is also
    appended to this text file (in addition to being passed to log_fn),
    so a plain-text run record is kept on disk.

    log_fn: optional callback(message, tag) for progress/status logging
    (matches the (message, tag) signature used by the GUI's log console).
    """
    log_file_handle = None
    if log_path:
        log_path = Path(log_path)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_file_handle = open(log_path, "a", encoding="utf-8")
        log_file_handle.write(f"\n{'=' * 66}\nRun started: {pd.Timestamp.now():%Y-%m-%d %H:%M:%S}\n")

    def _log(message: str, tag: Optional[str] = "info") -> None:
        if log_fn:
            log_fn(message, tag)
        if log_file_handle:
            log_file_handle.write(message + "\n")
            log_file_handle.flush()

    try:
        merged = load_and_merge(input_paths, sheet_name=sheet_name, header_row=header_row, log_fn=_log)
        rows_merged = len(merged)

        columns = list(merged.columns)
        lcp_col = lcp_col or _guess_column(columns, _LCP_PATTERNS)
        nap_col = nap_col or _guess_column(columns, _NAP_PATTERNS)
        area_col = area_col or _guess_column(columns, _AREA_PATTERNS)
        ports_col = ports_col or _guess_column(columns, _PORTS_PATTERNS)

        missing = [
            name for name, val in
            [("LCP", lcp_col), ("NAP", nap_col), ("AREA", area_col), ("PORTS", ports_col)]
            if val is None
        ]
        if missing:
            raise ValueError(
                f"Could not detect column(s) for: {', '.join(missing)}. "
                f"Available columns: {columns}. Pass the column name(s) explicitly "
                f"(e.g. lcp_col='LCP') to fix this."
            )

        non_blank, blank_rows_removed = remove_blank_rows(merged)
        _log(f"Removed {blank_rows_removed} fully blank row(s).", "info")

        deduped, duplicates_removed, duplicate_rows = remove_full_row_duplicates(non_blank)
        _log(f"Removed {duplicates_removed} exact-duplicate row(s).", "info")

        sorted_df = sort_by_lcp_nap(deduped, lcp_col, nap_col)
        pivot_df, grand_total = build_area_pivot(sorted_df, area_col, ports_col)
        _log(f"Built pivot: {len(pivot_df) - 1} area(s), grand total {grand_total} ports.", "info")

        result = MergeResult(
            combined_df=sorted_df,
            pivot_df=pivot_df,
            duplicate_rows_df=duplicate_rows,
            rows_merged=rows_merged,
            blank_rows_removed=blank_rows_removed,
            duplicates_removed=duplicates_removed,
            rows_final=len(sorted_df),
            grand_total_ports=grand_total,
            columns_used={"lcp": lcp_col, "nap": nap_col, "area": area_col, "ports": ports_col},
        )

        if output_path:
            output_path = Path(output_path)
            write_formatted_workbook(sorted_df, pivot_df, output_path, combined_sheet_name, pivot_sheet_name)
            result.output_path = output_path
            _log(f"Workbook saved to: {output_path}", "ok")

            if duplicates_csv_path is None and write_duplicates_csv:
                duplicates_csv_path = output_path.with_name(output_path.stem + "_duplicates_removed.csv")

        if duplicates_csv_path and len(duplicate_rows) > 0:
            written = write_removed_rows_csv(duplicate_rows, duplicates_csv_path)
            result.duplicates_csv_path = written
            _log(f"Duplicate rows ({len(duplicate_rows)}) saved to: {written}", "ok")
        elif duplicates_csv_path:
            _log("No duplicate rows found -- duplicates CSV was not created.", "info")

        if log_path:
            result.log_path = log_path

        return result
    finally:
        if log_file_handle:
            log_file_handle.close()


# --------------------------------------------------------------------------- #
# CLI entry point (optional convenience -- delete if you only need the API)
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Merge, dedupe, sort, and pivot Excel sheets.")
    parser.add_argument("inputs", nargs="+", help="Input .xlsx files")
    parser.add_argument("-o", "--output", default="merged_network_data.xlsx", help="Output .xlsx path")
    parser.add_argument("--header-row", type=int, default=0,
                         help="0-based row index of the real column headers (use 1 if there's a hidden field-name row above the display header)")
    parser.add_argument("--lcp-col", default=None)
    parser.add_argument("--nap-col", default=None)
    parser.add_argument("--area-col", default=None)
    parser.add_argument("--ports-col", default=None)
    parser.add_argument("--duplicates-csv", default=None,
                         help="Where to save removed duplicate rows (default: <output>_duplicates_removed.csv)")
    parser.add_argument("--log-file", default=None,
                         help="Also append the run log to this text file")
    args = parser.parse_args()

    res = process_workbooks(
        input_paths=args.inputs,
        output_path=args.output,
        header_row=args.header_row,
        lcp_col=args.lcp_col,
        nap_col=args.nap_col,
        area_col=args.area_col,
        ports_col=args.ports_col,
        duplicates_csv_path=args.duplicates_csv,
        log_path=args.log_file,
        log_fn=lambda msg, tag=None: print(msg),
    )
    print(res.summary())
